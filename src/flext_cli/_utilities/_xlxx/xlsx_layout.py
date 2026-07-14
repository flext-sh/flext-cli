"""Apply typed worksheet layout plans through public openpyxl APIs."""

from __future__ import annotations

from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from flext_cli import c, m, p, r, t

from .xlsx_addresses import FlextCliUtilitiesXlsxAddresses


class FlextCliUtilitiesXlsxLayout(FlextCliUtilitiesXlsxAddresses):
    """Apply comments, links, dimensions, grouping, views, and merges."""

    # mro-wkii.17.26 (xlsx-a): preserve typed results across cohesive phases.
    @classmethod
    def _apply_layout_annotations(
        cls, worksheet: Worksheet, plan: m.Cli.XlsxSheetLayoutPlan
    ) -> p.Result[bool]:
        for item in plan.comments:
            comment = Comment(item.text, item.author)
            if item.width is not None:
                comment.width = item.width
            if item.height is not None:
                comment.height = item.height
            cell = worksheet.cell(item.at.row, item.at.column)
            if not isinstance(cell, Cell):
                return r[bool].fail(
                    f"Cannot comment merged cell: row={item.at.row}, "
                    f"column={item.at.column}"
                )
            cell.comment = comment
        for item in plan.hyperlinks:
            cell = worksheet.cell(item.at.row, item.at.column)
            if not isinstance(cell, Cell):
                return r[bool].fail(
                    f"Cannot link merged cell: row={item.at.row}, "
                    f"column={item.at.column}"
                )
            if item.kind == "external":
                cell.hyperlink = item.target
            else:
                destination = cls._cell_ref(item.destination)
                sheet = cls._sheet_ref(item.destination_sheet)
                cell.hyperlink = f"#{sheet}!{destination}"
        return r[bool].ok(True)

    @classmethod
    def _apply_layout_structure(
        cls, worksheet: Worksheet, plan: m.Cli.XlsxSheetLayoutPlan
    ) -> p.Result[bool]:
        for item in plan.dimensions:
            for index in range(item.first, item.last + 1):
                if item.axis == "row":
                    dimension = worksheet.row_dimensions[index]
                    dimension.height = item.size
                else:
                    dimension = worksheet.column_dimensions[cls._column_ref(index)]
                    if item.size is not None:
                        dimension.width = item.size
                dimension.hidden = item.hidden
        for item in plan.groups:
            if item.axis == "row":
                worksheet.row_dimensions.group(
                    item.first,
                    item.last,
                    outline_level=item.outline_level,
                    hidden=item.hidden,
                )
            else:
                worksheet.column_dimensions.group(
                    cls._column_ref(item.first),
                    cls._column_ref(item.last),
                    outline_level=item.outline_level,
                    hidden=item.hidden,
                )
        return r[bool].ok(True)

    @classmethod
    def _apply_layout_view(
        cls, worksheet: Worksheet, plan: m.Cli.XlsxSheetLayoutPlan
    ) -> p.Result[bool]:
        if plan.freeze_pane is not None:
            worksheet.freeze_panes = cls._cell_ref(plan.freeze_pane.at)
        if plan.auto_filter is not None:
            worksheet.auto_filter.ref = cls._range_ref(plan.auto_filter.area)
        if plan.view is not None:
            worksheet.sheet_state = plan.view.visibility
            worksheet.sheet_properties.tabColor = plan.view.tab_color
        return r[bool].ok(True)

    @classmethod
    def _apply_layout_merges(
        cls, worksheet: Worksheet, plan: m.Cli.XlsxSheetLayoutPlan
    ) -> p.Result[bool]:
        for item in plan.merges:
            worksheet.merge_cells(cls._range_ref(item.area))
        return r[bool].ok(True)

    # NOTE (multi-agent, mro-j2yt.1): merge operations run last so comments
    # and hyperlinks can still target every concrete cell in the plan.
    @classmethod
    def _apply_layout(
        cls, worksheet: Worksheet, plan: m.Cli.XlsxSheetLayoutPlan
    ) -> p.Result[bool]:
        phases: tuple[t.Cli.NullaryOperation[p.Result[bool]], ...] = (
            lambda: cls._apply_layout_annotations(worksheet, plan),
            lambda: cls._apply_layout_structure(worksheet, plan),
            lambda: cls._apply_layout_view(worksheet, plan),
            lambda: cls._apply_layout_merges(worksheet, plan),
        )
        for phase in phases:
            try:
                result: p.Result[bool] = phase()
            except c.EXC_ATTR_KEY_TYPE_VALUE as exc:
                detail = str(exc).strip() or exc.__class__.__name__
                return r[bool].fail(
                    f"{c.Cli.XlsxError.RENDER_FAILED}: {detail}", exception=exc
                )
            if result.failure:
                return result
        return r[bool].ok(True)


__all__: tuple[str, ...] = ("FlextCliUtilitiesXlsxLayout",)
