"""Model-driven XLSX rendering contract tests."""

from __future__ import annotations

import datetime as dt
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

from flext_cli import cli, m, p
from flext_tests import tm


def _defined_name_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = 17
    sheet["A2"] = "second"
    workbook.defined_names.add(DefinedName("ScalarValue", attr_text="'Data'!$A$1"))
    workbook.defined_names.add(DefinedName("RangeValues", attr_text="'Data'!$A$1:$A$2"))
    workbook.defined_names.add(DefinedName("MalformedValue", attr_text='"constant"'))
    target = BytesIO()
    workbook.save(target)
    return target.getvalue()


def test_xlsx_defined_name_values_resolves_cached_scalar() -> None:
    source = _defined_name_workbook()

    result = cli.xlsx_defined_name_values(
        m.Cli.XlsxDefinedNameValuesRequest(source=source, name="ScalarValue")
    )

    tm.that(result.success, eq=True, msg=result.error)
    tm.that(result.value.name, eq="ScalarValue")
    tm.that(len(result.value.cells), eq=1)
    tm.that(result.value.cells[0].sheet, eq="Data")
    tm.that(result.value.cells[0].coordinate, eq="A1")
    tm.that(result.value.cells[0].value.kind, eq="integer")
    tm.that(result.value.cells[0].value.value, eq=17)


def test_xlsx_defined_name_values_resolves_cached_range() -> None:
    source = _defined_name_workbook()

    result = cli.xlsx_defined_name_values(
        m.Cli.XlsxDefinedNameValuesRequest(source=source, name="RangeValues")
    )

    tm.that(result.success, eq=True, msg=result.error)
    tm.that(
        tuple((cell.coordinate, cell.value.kind) for cell in result.value.cells),
        eq=(("A1", "integer"), ("A2", "text")),
    )


def test_xlsx_defined_name_values_fails_when_name_is_missing() -> None:
    source = _defined_name_workbook()

    result = cli.xlsx_defined_name_values(
        m.Cli.XlsxDefinedNameValuesRequest(source=source, name="MissingValue")
    )

    tm.that(result.success, eq=False)
    tm.that(result.error, has="xlsx_defined_name_missing")


def test_xlsx_defined_name_values_fails_for_non_range_name() -> None:
    source = _defined_name_workbook()

    result = cli.xlsx_defined_name_values(
        m.Cli.XlsxDefinedNameValuesRequest(source=source, name="MalformedValue")
    )

    tm.that(result.success, eq=False)
    tm.that(result.error, has="xlsx_defined_name_invalid")


def test_xlsx_datetime_rejects_unrepresentable_timezone() -> None:
    """XLSX ingress fails before vendor serialization for aware datetimes."""
    with pytest.raises(ValueError, match="Input should not have timezone info"):
        m.Cli.XlsxDateTimeValue(value=dt.datetime(2026, 7, 13, tzinfo=dt.UTC))


def test_xlsx_render_executes_typed_runtime_plan() -> None:
    """One immutable plan owns formulas, rules, tables, names, and protection."""
    source_workbook = Workbook()
    source_cell = source_workbook.worksheets[0]["A1"]
    source_cell.value = "visual"
    source_cell.font = Font(name="Aptos", bold=True, color="FFFFFFFF")
    source = BytesIO()
    source_workbook.save(source)
    template_result = cli.xlsx_style_template(
        m.Cli.XlsxStyleTemplateRequest(
            source=source.getvalue(), style_name_prefix="runtime"
        )
    )
    tm.that(template_result.success, eq=True)
    style_name = template_result.value.style_map[0].style_name
    data_area = m.Cli.XlsxCellRange(
        first=m.Cli.XlsxCellAddress(row=1, column=1),
        last=m.Cli.XlsxCellAddress(row=2, column=2),
    )
    value_area = m.Cli.XlsxCellRange(
        first=m.Cli.XlsxCellAddress(row=2, column=2),
        last=m.Cli.XlsxCellAddress(row=2, column=2),
    )
    plan = m.Cli.XlsxWorkbookPlan(
        sheets=(
            m.Cli.XlsxSheetPlan(
                name="Data",
                cells=(
                    m.Cli.XlsxCellPlan(
                        at=m.Cli.XlsxCellAddress(row=1, column=1),
                        value=m.Cli.XlsxTextValue(value="Item"),
                        style=style_name,
                    ),
                    m.Cli.XlsxCellPlan(
                        at=m.Cli.XlsxCellAddress(row=1, column=2),
                        value=m.Cli.XlsxTextValue(value="Total"),
                        style=style_name,
                    ),
                    m.Cli.XlsxCellPlan(
                        at=m.Cli.XlsxCellAddress(row=2, column=1),
                        value=m.Cli.XlsxTextValue(value="Example"),
                        style=style_name,
                    ),
                    m.Cli.XlsxCellPlan(
                        at=m.Cli.XlsxCellAddress(row=2, column=2),
                        value=m.Cli.XlsxFormulaValue(value="=1+1"),
                        style=style_name,
                    ),
                ),
                tables=(
                    m.Cli.XlsxTablePlan(
                        name="DataTable", area=data_area, style="TableStyleMedium2"
                    ),
                ),
                layout=m.Cli.XlsxSheetLayoutPlan(
                    freeze_pane=m.Cli.XlsxFreezePanePlan(
                        at=m.Cli.XlsxCellAddress(row=2, column=1)
                    )
                ),
                rules=m.Cli.XlsxSheetRulesPlan(
                    validations=(
                        m.Cli.XlsxListValidationPlan(
                            area=value_area,
                            source=m.Cli.XlsxInlineListSource(values=("1", "2")),
                            messages=m.Cli.XlsxValidationMessages(),
                        ),
                    ),
                    conditional_formats=(
                        m.Cli.XlsxFormulaFormatPlan(
                            area=value_area, expressions=("B2>1",), style=style_name
                        ),
                    ),
                    protection=m.Cli.XlsxSheetProtectionPlan(
                        credential=m.Cli.XlsxPlainProtectionCredential(value="secret"),
                        permissions=m.Cli.XlsxProtectionPermissions(
                            allow_select_unlocked=True
                        ),
                        cells=(
                            m.Cli.XlsxCellProtectionPlan(
                                area=value_area, locked=True, hidden=True
                            ),
                        ),
                    ),
                ),
            ),
            m.Cli.XlsxSheetPlan(
                name="Summary",
                cells=(),
                tables=(),
                layout=m.Cli.XlsxSheetLayoutPlan(),
                rules=m.Cli.XlsxSheetRulesPlan(),
            ),
        ),
        defined_names=(
            m.Cli.XlsxRangeDefinedNamePlan(
                name="DataRange", sheet="Data", area=data_area
            ),
        ),
    )

    result = cli.xlsx_render(
        m.Cli.XlsxRenderRequest(template=template_result.value.content, plan=plan)
    )

    tm.that(result.success, eq=True)
    tm.that(result.value.plan is plan, eq=True)
    rendered = load_workbook(BytesIO(result.value.content), data_only=False)
    tm.that(tuple(rendered.sheetnames), eq=("Data", "Summary"))
    data = rendered.worksheets[0]
    tm.that(data["B2"].value, eq="=1+1")
    tm.that(data["B2"].style, eq=style_name)
    tm.that(data["B2"].protection.locked, eq=True)
    tm.that(data["B2"].protection.hidden, eq=True)
    tm.that(data.protection.sheet, eq=True)
    tm.that("DataTable" in data.tables, eq=True)
    tm.that(len(data.data_validations.dataValidation), eq=1)
    tm.that(len(data.conditional_formatting), eq=1)
    tm.that("DataRange" in rendered.defined_names, eq=True)
    tm.that(rendered.calculation.fullCalcOnLoad, eq=True)

    snapshot = cli.xlsx_snapshot(
        m.Cli.XlsxSnapshotRequest(source=result.value.content, data_only=False)
    )
    cached = cli.xlsx_snapshot(
        m.Cli.XlsxSnapshotRequest(source=result.value.content, data_only=True)
    )

    tm.that(snapshot.success, eq=True, msg=snapshot.error)
    tm.that(cached.success, eq=True, msg=cached.error)
    tm.that(isinstance(snapshot.value, p.Cli.XlsxWorkbookSnapshot), eq=True)
    tm.that(tuple(item.name for item in snapshot.value.sheets), eq=("Data", "Summary"))
    data_snapshot = snapshot.value.sheets[0]
    formula_cell = next(item for item in data_snapshot.cells if item.coordinate == "B2")
    cached_formula_cell = next(
        item for item in cached.value.sheets[0].cells if item.coordinate == "B2"
    )
    tm.that(formula_cell.value.kind, eq="formula")
    tm.that(formula_cell.formula, eq="=1+1")
    tm.that(formula_cell.style_name, eq=style_name)
    tm.that(cached_formula_cell.value.kind, eq="blank")
    tm.that(cached_formula_cell.formula, eq="=1+1")
    tm.that(snapshot.value.formula_count, eq=1)
    tm.that(snapshot.value.literal_count, eq=3)
    tm.that(cached.value.formula_count, eq=1)
    tm.that(data_snapshot.data_validation_count, eq=1)
    tm.that(data_snapshot.conditional_format_count, eq=1)
    tm.that(data_snapshot.protection.enabled, eq=True)
    tm.that(data_snapshot.protection.legacy_password_hash, eq=data.protection.password)
    tm.that(data_snapshot.tables[0].reference, eq="A1:B2")
    tm.that(snapshot.value.defined_names[0].name, eq="DataRange")
