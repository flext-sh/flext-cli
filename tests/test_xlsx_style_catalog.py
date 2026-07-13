"""Format-only XLSX template contract tests."""

from __future__ import annotations

from io import BytesIO

from flext_tests import tm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection

from flext_cli import cli, m

_SOURCE_STYLE_COUNT = 2


def test_xlsx_style_template_deduplicates_visuals_and_discards_content() -> None:
    """Protection differences never split a visual style signature."""
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    for coordinate, locked in (("A1", True), ("B1", False)):
        cell = worksheet[coordinate]
        cell.value = coordinate
        cell.font = Font(name="Aptos", size=12, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FF112233")
        cell.protection = Protection(locked=locked, hidden=not locked)
    source = BytesIO()
    workbook.save(source)

    result = cli.xlsx_style_template(
        m.Cli.XlsxStyleTemplateRequest(
            source=source.getvalue(), style_name_prefix="visual"
        )
    )

    tm.that(result.success, eq=True)
    tm.that(len(result.value.style_map), eq=_SOURCE_STYLE_COUNT)
    style_names = frozenset(entry.style_name for entry in result.value.style_map)
    tm.that(len(style_names), eq=1)
    tm.that(next(iter(style_names)).startswith("visual_"), eq=True)
    template = load_workbook(BytesIO(result.value.content))
    tm.that(tuple(template.named_styles), eq=("Normal", next(iter(style_names))))
    worksheet = template.worksheets[0]
    tm.that(worksheet.max_row, eq=1)
    tm.that(worksheet.max_column, eq=1)
    tm.that(worksheet["A1"].value, eq=None)

    inspection = cli.xlsx_inspect(
        m.Cli.XlsxArchiveInspectionRequest(
            source=result.value.content,
            policy=m.Cli.XlsxArchivePolicy(
                forbidden_worksheet_tags=frozenset(("c", "f", "mergeCells")),
                required_worksheet_count=1,
                reject_defined_names=True,
                reject_style_protection=True,
            ),
        )
    )
    tm.that(inspection.success, eq=True)
    tm.that(inspection.value.clean, eq=True)
